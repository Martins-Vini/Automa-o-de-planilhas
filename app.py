from openpyxl import load_workbook, Workbook; #type: ignore

planilhas = load_workbook("planilhaTeste.xlsx");
planilhaAlunos = planilhas['tabelaAlunos'];

# Iteração sobre dados da planilha e pegando os valores das células;

for linha in planilhaAlunos.iter_rows(values_only=True):
    print(linha);
    
planilhaOutrosAlunos = Workbook()
pagina1 = planilhaOutrosAlunos.active
pagina1.title = "tabelaOutrosAlunos"

with open("Alunos.txt", "r", encoding="utf-8") as arquivo:
    linhas = arquivo.readlines()
    for linha in linhas:
        dados = linha.strip().split(",")
        pagina1.append(dados)
        
planilhaOutrosAlunos.save("planilhaOutrosAlunos.xlsx");